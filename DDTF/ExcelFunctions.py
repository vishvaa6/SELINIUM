from openpyxl import load_workbook


class VishvaaExcelFunctions:


   def __init__(self, file_name, sheet_name):
       self.file = file_name
       self.sheet = sheet_name


   # Fetch the total row count from Excel sheet
   def row_count(self):
       workbook = load_workbook(self.file)
       sheet = workbook[self.sheet]
       return sheet.max_row


   # Fetch the total column count from the Excel sheet
   def column_count(self):
       workbook = load_workbook(self.file)
       sheet = workbook[self.sheet]
       return sheet.max_column


   # Read the data from Excel sheet of specific Row and Column
   def read_data(self, row_number, column_number):
       workbook = load_workbook(self.file)
       sheet = workbook[self.sheet]
       return sheet.cell(row=row_number, column=column_number).value


   # Write the data into an Excel sheet on a specific Row and Column
   def write_data(self, row_number, column_number, data):
       workbook = load_workbook(self.file)
       sheet = workbook[self.sheet]
       sheet.cell(row=row_number, column=column_number).value = data
       workbook.save(self.file)
